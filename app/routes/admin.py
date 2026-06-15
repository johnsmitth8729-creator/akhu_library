from datetime import datetime, timedelta
from io import BytesIO, StringIO
import csv

from flask import Blueprint, render_template, request, flash, redirect, url_for, send_file, session, make_response, current_app, abort
from flask_login import login_required, current_user
from sqlalchemy import desc, func, or_

from app import db
from app.models.user import User
from app.models.book import Book, PhysicalBook, Category, ReadingProgress, PDFBookmark
from app.models.borrow import BorrowRequest, BorrowHistory, FavoriteBook
from app.models.competition import ReadingCompetition, QuizAttempt, CompetitionCertificate
from app.models.activity import ActivityLog
from app.models.system import Settings, Notification, Review, SiteBanner, Announcement
from app.forms.admin_forms import AdminCreateUserForm, AdminEditUserForm
from app.forms.faculty_forms import FacultyForm
from app.models.faculty import Faculty
from app.services.user_creation import assign_faculty
from app.services.user_import import default_temp_password
from app.utils.datetime import now_local
from app.utils.helpers import log_activity
from app.utils.faculty_helpers import load_faculty_choices
from app.utils.phone import normalize_phone
from app.services.analytics import (
    build_admin_dashboard_context,
    build_admin_statistics_context,
)


admin_bp = Blueprint("admin", __name__)


@admin_bp.before_request
@login_required
def require_admin():
    if not current_user.is_admin:
        flash("Admin access required.", "danger")
        return redirect(url_for("main.home"))


@admin_bp.route("/dashboard")
def dashboard():
    return render_template(
        "admin/dashboard.html",
        **build_admin_dashboard_context(),
    )


@admin_bp.route("/statistics")
def statistics():
    context = build_admin_statistics_context()

    return render_template(
        "admin/statistics.html",
        **context
    )


def _filter_rows(rows, q, faculty, group):
    if q:
        q = q.lower()
        rows = [
            row
            for row in rows
            if q in str(row.get("name", "")).lower()
            or q in str(row.get("faculty", "")).lower()
            or q in str(row.get("group", "")).lower()
            or q in str(row.get("title", "")).lower()
        ]
    if faculty:
        rows = [row for row in rows if str(row.get("faculty", "")).lower() == faculty.lower()]
    if group:
        rows = [row for row in rows if str(row.get("group", "")).lower() == group.lower()]
    return rows


def _download_csv(filename, columns, rows):
    csv_io = StringIO()
    writer = csv.writer(csv_io)
    writer.writerow([col["label"] for col in columns])
    for row in rows:
        writer.writerow([row.get(col["key"], "") for col in columns])
    response = make_response(csv_io.getvalue())
    response.headers["Content-Type"] = "text/csv; charset=utf-8"
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response


def _download_excel(filename, columns, rows):
    try:
        from openpyxl import Workbook
    except ImportError:
        flash("Excel export requires openpyxl. Install requirements and try again.", "danger")
        return redirect(request.path)

    workbook = Workbook()
    sheet = workbook.active
    sheet.append([col["label"] for col in columns])
    for row in rows:
        sheet.append([row.get(col["key"], "") for col in columns])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _sort_rows(rows, sort_key, sort_dir):
    if not sort_key:
        return rows

    direction = sort_dir.lower()
    reverse = direction == "desc"

    def sort_key_fn(row):
        value = row.get(sort_key, "")

        if isinstance(value, str):
            return value.lower()
        if value is None:
            return ""
        return value

    try:
        return sorted(rows, key=sort_key_fn, reverse=reverse)
    except TypeError:
        return sorted(rows, key=lambda row: str(row.get(sort_key, "")).lower(), reverse=reverse)


def _paginate_rows(rows, page, per_page=20):
    total = len(rows)
    pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, pages))
    start = (page - 1) * per_page
    end = start + per_page
    items = rows[start:end]

    return {
        "items": items,
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": pages,
        "has_prev": page > 1,
        "has_next": page < pages,
        "prev_num": page - 1,
        "next_num": page + 1,
        "start_index": start + 1 if total else 0,
        "end_index": min(total, end),
    }


@admin_bp.route("/statistics/readers")
def statistics_readers():
    context = build_admin_statistics_context()
    rows = context["top_readers_overall"]
    q = request.args.get("q", "").strip()
    faculty = request.args.get("faculty", "").strip()
    group = request.args.get("group", "").strip()
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 20, type=int)
    sort = request.args.get("sort", "score").strip()
    direction = request.args.get("direction", "desc").strip().lower()

    rows = _filter_rows(rows, q, faculty, group)
    rows = _sort_rows(rows, sort, direction)

    export = request.args.get("export", "")
    if export == "csv":
        return _download_csv("top_readers.csv", columns, rows)
    if export == "excel":
        return _download_excel("top_readers.xlsx", columns, rows)

    pagination = _paginate_rows(rows, page, per_page)
    rows = pagination["items"]

    columns = [
        {"key": "name", "label": "User"},
        {"key": "faculty", "label": "Faculty"},
        {"key": "group", "label": "Group"},
        {"key": "score", "label": "Activity Score"},
        {"key": "borrows", "label": "Borrows"},
        {"key": "pages_read", "label": "Pages Read"},
        {"key": "favorites", "label": "Favorites"},
        {"key": "reviews", "label": "Reviews"},
        {"key": "competitions_joined", "label": "Competitions"},
        {"key": "last_active", "label": "Last Active"},
    ]

    return render_template(
        "admin/statistics_module.html",
        page_title="Top Readers Analytics",
        description="Identify the most active readers, based on borrow behavior, digital reading progress, favorites, and competition participation.",
        rows=rows,
        columns=columns,
        pagination=pagination,
        filters={
            "q": q,
            "faculty": faculty,
            "group": group,
            "sort": sort,
            "direction": direction,
            "page": page,
            "per_page": per_page,
            "faculties": Faculty.query.order_by(Faculty.name.asc()).all(),
            "groups": sorted({row["group"] for row in context["top_readers_overall"] if row["group"] and row["group"] != "—"}),
        },
        export_url="admin.statistics_readers",
    )


@admin_bp.route("/statistics/borrowers")
def statistics_borrowers():
    context = build_admin_statistics_context()
    rows = context["top_borrowers_overall"]
    q = request.args.get("q", "").strip()
    faculty = request.args.get("faculty", "").strip()
    group = request.args.get("group", "").strip()
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 20, type=int)
    sort = request.args.get("sort", "borrow_score").strip()
    direction = request.args.get("direction", "desc").strip().lower()

    rows = _filter_rows(rows, q, faculty, group)
    rows = _sort_rows(rows, sort, direction)

    columns = [
        {"key": "name", "label": "User"},
        {"key": "faculty", "label": "Faculty"},
        {"key": "group", "label": "Group"},
        {"key": "approved_requests", "label": "Approved Requests"},
        {"key": "returned_books", "label": "Returned Books"},
        {"key": "active_borrowings", "label": "Active Borrowings"},
        {"key": "overdue_books", "label": "Overdue Books"},
        {"key": "borrow_score", "label": "Borrow Score"},
    ]

    export = request.args.get("export", "")
    if export == "csv":
        return _download_csv("top_borrowers.csv", columns, rows)
    if export == "excel":
        return _download_excel("top_borrowers.xlsx", columns, rows)

    pagination = _paginate_rows(rows, page, per_page)
    rows = pagination["items"]

    return render_template(
        "admin/statistics_module.html",
        page_title="Top Borrowers Analytics",
        description="Track patrons who borrow the most physical books and maintain the strongest borrowing records.",
        rows=rows,
        columns=columns,
        pagination=pagination,
        filters={
            "q": q,
            "faculty": faculty,
            "group": group,
            "sort": sort,
            "direction": direction,
            "page": page,
            "per_page": per_page,
            "faculties": Faculty.query.order_by(Faculty.name.asc()).all(),
            "groups": sorted({row["group"] for row in context["top_borrowers_overall"] if row["group"] and row["group"] != "—"}),
        },
        export_url="admin.statistics_borrowers",
    )


@admin_bp.route("/statistics/competition-leaders")
def statistics_competition_leaders():
    context = build_admin_statistics_context()
    rows = context["top_competition_leaders"]
    q = request.args.get("q", "").strip()
    faculty = request.args.get("faculty", "").strip()
    group = request.args.get("group", "").strip()
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 20, type=int)
    sort = request.args.get("sort", "competition_score").strip()
    direction = request.args.get("direction", "desc").strip().lower()

    rows = _filter_rows(rows, q, faculty, group)
    rows = _sort_rows(rows, sort, direction)

    columns = [
        {"key": "name", "label": "User"},
        {"key": "faculty", "label": "Faculty"},
        {"key": "group", "label": "Group"},
        {"key": "competitions_joined", "label": "Competitions Joined"},
        {"key": "certificates", "label": "Certificates"},
        {"key": "bronze_medals", "label": "Bronze Medals"},
        {"key": "silver_medals", "label": "Silver Medals"},
        {"key": "gold_medals", "label": "Gold Medals"},
        {"key": "champion_positions", "label": "Champion Positions"},
        {"key": "competition_score", "label": "Competition Score"},
    ]

    export = request.args.get("export", "")
    if export == "csv":
        return _download_csv("competition_leaders.csv", columns, rows)
    if export == "excel":
        return _download_excel("competition_leaders.xlsx", columns, rows)

    pagination = _paginate_rows(rows, page, per_page)
    rows = pagination["items"]

    return render_template(
        "admin/statistics_module.html",
        page_title="Competition Leaders Analytics",
        description="Analyze top competition performers by participation, certificates, medals, and top finishes.",
        rows=rows,
        columns=columns,
        pagination=pagination,
        filters={
            "q": q,
            "faculty": faculty,
            "group": group,
            "sort": sort,
            "direction": direction,
            "page": page,
            "per_page": per_page,
            "faculties": Faculty.query.order_by(Faculty.name.asc()).all(),
            "groups": sorted({row["group"] for row in context["top_competition_leaders"] if row["group"] and row["group"] != "—"}),
        },
        export_url="admin.statistics_competition_leaders",
    )


@admin_bp.route("/statistics/rankings")
def statistics_rankings():
    context = build_admin_statistics_context()
    q = request.args.get("q", "").strip()
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 20, type=int)
    sort = request.args.get("sort", "score").strip()
    direction = request.args.get("direction", "desc").strip().lower()

    faculty_rows = _sort_rows(_filter_rows(context["faculty_rankings"], q, "", ""), sort, direction)
    group_rows = _sort_rows(_filter_rows(context["group_rankings"], q, "", ""), sort, direction)

    faculty_pagination = _paginate_rows(faculty_rows, page, per_page)
    group_pagination = _paginate_rows(group_rows, page, per_page)

    return render_template(
        "admin/statistics_rankings.html",
        page_title="Faculty & Group Rankings",
        description="Compare faculty and group performance across reading, borrowing, and competition activity.",
        faculty_rankings=faculty_pagination["items"],
        group_rankings=group_pagination["items"],
        faculty_pagination=faculty_pagination,
        group_pagination=group_pagination,
        filters={
            "q": q,
            "sort": sort,
            "direction": direction,
            "page": page,
            "per_page": per_page,
        },
    )


@admin_bp.route("/statistics/physical-books")
def statistics_physical_books():
    context = build_admin_statistics_context()
    rows = context["popular_physical"]
    category_filter = request.args.get("category", "").strip()
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 20, type=int)
    sort = request.args.get("sort", "score").strip()
    direction = request.args.get("direction", "desc").strip().lower()

    if category_filter:
        rows = [row for row in rows if row.get("category", "").lower() == category_filter.lower()]

    columns = [
        {"key": "title", "label": "Book Title"},
        {"key": "author", "label": "Author"},
        {"key": "category", "label": "Category"},
        {"key": "borrows", "label": "Borrow Count"},
        {"key": "request_count", "label": "Request Count"},
        {"key": "rating", "label": "Rating"},
        {"key": "score", "label": "Popularity Score"},
    ]

    rows = _sort_rows(rows, sort, direction)
    export = request.args.get("export", "")
    if export == "csv":
        return _download_csv("popular_physical_books.csv", columns, rows)
    if export == "excel":
        return _download_excel("popular_physical_books.xlsx", columns, rows)

    pagination = _paginate_rows(rows, page, per_page)
    rows = pagination["items"]

    return render_template(
        "admin/statistics_module.html",
        page_title="Popular Physical Books Analytics",
        description="Find the most requested and borrowed physical books across the library.",
        rows=rows,
        columns=columns,
        pagination=pagination,
        filters={
            "category": category_filter,
            "categories": Category.query.order_by(Category.name.asc()).all(),
            "sort": sort,
            "direction": direction,
            "page": page,
            "per_page": per_page,
        },
        export_url="admin.statistics_physical_books",
    )


@admin_bp.route("/statistics/digital-books")
def statistics_digital_books():
    context = build_admin_statistics_context()
    rows = context["popular_digital"]
    category_filter = request.args.get("category", "").strip()
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 20, type=int)
    sort = request.args.get("sort", "score").strip()
    direction = request.args.get("direction", "desc").strip().lower()

    if category_filter:
        rows = [row for row in rows if row.get("category", "").lower() == category_filter.lower()]

    columns = [
        {"key": "title", "label": "Book Title"},
        {"key": "author", "label": "Author"},
        {"key": "category", "label": "Category"},
        {"key": "views", "label": "Views"},
        {"key": "reads", "label": "Readers"},
        {"key": "favorites", "label": "Favorites"},
        {"key": "reviews", "label": "Reviews"},
        {"key": "rating", "label": "Average Rating"},
        {"key": "score", "label": "Popularity Score"},
    ]

    rows = _sort_rows(rows, sort, direction)
    export = request.args.get("export", "")
    if export == "csv":
        return _download_csv("popular_digital_books.csv", columns, rows)
    if export == "excel":
        return _download_excel("popular_digital_books.xlsx", columns, rows)

    pagination = _paginate_rows(rows, page, per_page)
    rows = pagination["items"]

    return render_template(
        "admin/statistics_module.html",
        page_title="Popular Digital Books Analytics",
        description="Measure digital book engagement by views, readers, favorites, and reviews.",
        rows=rows,
        columns=columns,
        pagination=pagination,
        filters={
            "category": category_filter,
            "categories": Category.query.order_by(Category.name.asc()).all(),
            "sort": sort,
            "direction": direction,
            "page": page,
            "per_page": per_page,
        },
        export_url="admin.statistics_digital_books",
    )


@admin_bp.route("/users")
def manage_users():
    page = request.args.get("page", 1, type=int)
    q = request.args.get("q", "").strip()
    role = request.args.get("role", "").strip()
    status = request.args.get("status", "").strip()

    # Hide superadmin accounts from regular administrators
    query = User.query.filter(User.role != User.ROLE_SUPERADMIN)

    if q:
        query = query.filter(
            or_(
                User.username.ilike(f"%{q}%"),
                User.fullname.ilike(f"%{q}%"),
                User.email.ilike(f"%{q}%"),
                User.phone_number.ilike(f"%{q}%"),
                User.faculty.ilike(f"%{q}%"),
                User.group_name.ilike(f"%{q}%")
            )
        )

    if role in [User.ROLE_USER, User.ROLE_LIBRARIAN, User.ROLE_ADMIN]:
        query = query.filter(User.role == role)

    if status == "blocked":
        query = query.filter(User.is_blocked == True)
    elif status == "active":
        query = query.filter(User.is_blocked == False)

    pagination = query.order_by(
        desc(User.created_at)
    ).paginate(
        page=page,
        per_page=20,
        error_out=False
    )

    return render_template(
        "admin/users.html",
        pagination=pagination,
        q=q,
        role=role,
        status=status
    )


@admin_bp.route("/users/export")
def export_users():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash("Excel export requires openpyxl. Run pip install -r requirements.txt and try again.", "danger")
        return redirect(url_for("admin.manage_users"))

    q = request.args.get("q", "").strip()
    role = request.args.get("role", "").strip()
    status = request.args.get("status", "").strip()

    # Hide superadmin accounts from regular administrators
    query = User.query.filter(User.role != User.ROLE_SUPERADMIN)

    if q:
        query = query.filter(
            or_(
                User.username.ilike(f"%{q}%"),
                User.fullname.ilike(f"%{q}%"),
                User.email.ilike(f"%{q}%"),
                User.phone_number.ilike(f"%{q}%"),
                User.faculty.ilike(f"%{q}%"),
                User.group_name.ilike(f"%{q}%")
            )
        )

    if role in [User.ROLE_USER, User.ROLE_LIBRARIAN, User.ROLE_ADMIN]:
        query = query.filter(User.role == role)

    if status == "blocked":
        query = query.filter(User.is_blocked.is_(True))
    elif status == "active":
        query = query.filter(User.is_blocked.is_(False))

    users = query.order_by(desc(User.created_at)).all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Users"

    headers = [
        "ID", "Fullname", "Username", "Email", "Phone Number", "Role",
        "Faculty", "Group", "Status", "Email Verified",
        "Last Login", "Joined"
    ]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="0B3B6F")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D9E2F1"),
        right=Side(style="thin", color="D9E2F1"),
        top=Side(style="thin", color="D9E2F1"),
        bottom=Side(style="thin", color="D9E2F1")
    )

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for user in users:
        sheet.append([
            user.id,
            user.fullname,
            user.username,
            user.email,
            user.phone_number or "",
            user.role,
            user.faculty_display or "",
            user.group_name or "",
            "Blocked" if user.is_blocked else "Active",
            "Yes" if user.email_verified else "No",
            user.last_login_at.strftime("%Y-%m-%d %H:%M") if user.last_login_at else "",
            user.created_at.strftime("%Y-%m-%d %H:%M") if user.created_at else ""
        ])

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=12):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    widths = [8, 24, 18, 30, 18, 14, 24, 16, 14, 16, 20, 20]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="users_export.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# Placeholder bulk-action endpoint
@admin_bp.route("/users/bulk-action", methods=["POST"])
def bulk_user_action():
    """Handle bulk actions on selected users (placeholder)."""
    # Expect list of user IDs and an action name
    ids = request.form.getlist("ids")
    action = request.form.get("bulk_action")
    if not ids or not action:
        flash("No users selected or action missing.", "warning")
        return redirect(url_for("admin.manage_users"))
    # Placeholder logic – just flash a success message
    flash("Bulk action received successfully.", "success")
    return redirect(url_for("admin.manage_users"))


@admin_bp.route("/users/create", methods=["GET", "POST"])
def create_user():
    form = AdminCreateUserForm()
    load_faculty_choices(form)

    if form.validate_on_submit():
        email = form.email.data.strip().lower()
        username = form.username.data.strip()
        phone_raw = (form.phone_number.data or "").strip()
        phone_number = normalize_phone(phone_raw) if phone_raw else None
        role = form.role.data
        faculty_id = form.faculty_id.data or None
        group_name = (form.group_name.data or "").strip() or None

        if role != User.ROLE_USER:
            faculty_id = None
            group_name = None

        if phone_number and User.query.filter_by(phone_number=phone_number).first():
            flash("Phone number already in use.", "danger")
            return render_template("admin/user_form.html", form=form, title="Add User", mode="create")

        if User.query.filter(
            (User.email == email) | (User.username == username)
        ).first():
            flash("Email or username already in use.", "danger")
            return render_template("admin/user_form.html", form=form, title="Add User", mode="create")

        user = User(
            fullname=form.fullname.data.strip(),
            username=username,
            email=email,
            phone_number=phone_number,
            group_name=group_name,
            role=role
        )
        try:
            assign_faculty(user, faculty_id)
        except ValueError as exc:
            flash(str(exc), "danger")
            return render_template("admin/user_form.html", form=form, title="Add User", mode="create")

        user.set_password(form.password.data)

        db.session.add(user)
        db.session.commit()
        log_activity(current_user.id, f"Created {user.role} account for {user.username}")

        flash("User account created successfully.", "success")
        return redirect(url_for("admin.manage_users"))

    return render_template(
        "admin/user_form.html",
        form=form,
        title="Add User",
        mode="create",
        default_password_hint=default_temp_password(),
    )


@admin_bp.route("/users/<int:user_id>/edit", methods=["GET", "POST"])
def edit_user(user_id):
    user = User.query.get_or_404(user_id)
    if user.is_superadmin:
        abort(403)
    form = AdminEditUserForm(obj=user)
    load_faculty_choices(form)
    if request.method == "GET":
        form.faculty_id.data = user.faculty_id or 0

    if form.validate_on_submit():
        email = form.email.data.strip().lower()
        username = form.username.data.strip()
        phone_raw = (form.phone_number.data or "").strip()
        phone_number = normalize_phone(phone_raw) if phone_raw else None
        role = form.role.data
        faculty_id = form.faculty_id.data or None
        group_name = (form.group_name.data or "").strip() or None

        if role != User.ROLE_USER:
            faculty_id = None
            group_name = None

        duplicate = User.query.filter(
            User.id != user.id,
            (User.email == email) | (User.username == username)
        ).first()

        if duplicate:
            flash("Email or username already in use by another account.", "danger")
            return render_template("admin/user_form.html", form=form, title="Edit User", mode="edit", user=user)

        if phone_number:
            phone_duplicate = User.query.filter(
                User.id != user.id,
                User.phone_number == phone_number,
            ).first()
            if phone_duplicate:
                flash("Phone number already in use by another account.", "danger")
                return render_template(
                "admin/user_form.html",
                form=form,
                title="Edit User",
                mode="edit",
                user=user,
                default_password_hint=default_temp_password(),
            )

        if user.id == current_user.id and form.role.data != user.role:
            flash("You cannot change your own role.", "warning")
            return render_template(
                "admin/user_form.html",
                form=form,
                title="Edit User",
                mode="edit",
                user=user,
                default_password_hint=default_temp_password(),
            )

        user.fullname = form.fullname.data.strip()
        user.username = username
        user.email = email
        user.phone_number = phone_number
        user.group_name = group_name
        user.role = role

        try:
            assign_faculty(user, faculty_id)
        except ValueError as exc:
            flash(str(exc), "danger")
            return render_template(
                "admin/user_form.html",
                form=form,
                title="Edit User",
                mode="edit",
                user=user,
                default_password_hint=default_temp_password(),
            )

        new_password = None
        if form.password.data:
            user.set_password(form.password.data)
            new_password = form.password.data

        db.session.commit()
        log_activity(current_user.id, f"Updated user account {user.username}")

        if new_password:
            session[f"revealed_password_{user.id}"] = new_password
            flash("User updated. The new password is shown below — copy it before leaving this page.", "success")
            return redirect(url_for("admin.edit_user", user_id=user.id))

        flash("User information updated successfully.", "success")
        return redirect(url_for("admin.manage_users"))

    revealed_password = session.pop(f"revealed_password_{user.id}", None)

    return render_template(
        "admin/user_form.html",
        form=form,
        title="Edit User",
        mode="edit",
        user=user,
        revealed_password=revealed_password,
        default_password_hint=default_temp_password(),
    )


@admin_bp.route("/users/<int:user_id>/reset-password", methods=["POST"])
def reset_user_password(user_id):
    user = User.query.get_or_404(user_id)
    if user.is_superadmin:
        abort(403)

    if user.id == current_user.id:
        flash("Change your own password from your profile or account settings.", "warning")
        return redirect(url_for("admin.edit_user", user_id=user.id))

    new_password = default_temp_password()
    user.set_password(new_password)
    db.session.commit()
    log_activity(current_user.id, f"Reset password for user {user.username}")

    session[f"revealed_password_{user.id}"] = new_password
    flash("Password reset to the default temporary password. Copy it from the box below.", "success")
    return redirect(url_for("admin.edit_user", user_id=user.id))


@admin_bp.route("/users/<int:user_id>/role", methods=["POST"])
def user_role(user_id):
    user = User.query.get_or_404(user_id)
    if user.is_superadmin:
        abort(403)

    if user.id == current_user.id:
        flash("You cannot change your own role.", "warning")
        return redirect(url_for("admin.manage_users"))

    new_role = request.form.get("role")

    if new_role not in [
        User.ROLE_USER,
        User.ROLE_LIBRARIAN,
        User.ROLE_ADMIN
    ]:
        flash("Invalid role specified.", "danger")
        return redirect(url_for("admin.manage_users"))

    user.role = new_role
    if new_role != User.ROLE_USER:
        user.faculty_id = None
        user.faculty = None
        user.group_name = None
    db.session.commit()

    flash(f"User role updated to {new_role}.", "success")
    return redirect(url_for("admin.manage_users"))


@admin_bp.route("/users/<int:user_id>/block", methods=["POST"])
def block_user(user_id):
    user = User.query.get_or_404(user_id)
    if user.is_superadmin:
        abort(403)

    if user.id == current_user.id:
        flash("You cannot block yourself.", "warning")
        return redirect(url_for("admin.manage_users"))

    if user.is_admin:
        flash("Admin users cannot be blocked from this panel.", "danger")
        return redirect(url_for("admin.manage_users"))

    user.is_blocked = True
    db.session.commit()

    flash(f"User {user.username} has been blocked.", "success")
    return redirect(url_for("admin.manage_users"))


@admin_bp.route("/users/<int:user_id>/unblock", methods=["POST"])
def unblock_user(user_id):
    user = User.query.get_or_404(user_id)
    if user.is_superadmin:
        abort(403)

    user.is_blocked = False
    db.session.commit()

    flash(f"User {user.username} has been unblocked.", "success")
    return redirect(url_for("admin.manage_users"))


@admin_bp.route("/users/<int:user_id>/delete", methods=["POST"])
def delete_user(user_id):
    user = User.query.get_or_404(user_id)
    if user.is_superadmin:
        abort(403)

    if user.id == current_user.id:
        flash("You cannot delete yourself.", "warning")
        return redirect(url_for("admin.manage_users"))

    if user.is_admin:
        flash("Admin users cannot be deleted from this panel.", "danger")
        return redirect(url_for("admin.manage_users"))

    active_borrows = BorrowHistory.query.filter(
        BorrowHistory.user_id == user.id,
        BorrowHistory.status.in_(["Borrowed", "Overdue"])
    ).count()

    if active_borrows:
        flash("This user cannot be deleted while they have active borrowed books.", "danger")
        return redirect(url_for("admin.manage_users"))

    BorrowRequest.query.filter_by(user_id=user.id).delete(synchronize_session=False)
    BorrowHistory.query.filter_by(user_id=user.id).delete(synchronize_session=False)
    FavoriteBook.query.filter_by(user_id=user.id).delete(synchronize_session=False)
    ReadingProgress.query.filter_by(user_id=user.id).delete(synchronize_session=False)
    PDFBookmark.query.filter_by(user_id=user.id).delete(synchronize_session=False)
    Notification.query.filter_by(user_id=user.id).delete(synchronize_session=False)
    Review.query.filter_by(user_id=user.id).delete(synchronize_session=False)
    ActivityLog.query.filter_by(user_id=user.id).delete(synchronize_session=False)

    db.session.delete(user)
    db.session.commit()

    flash("User has been permanently deleted.", "success")
    return redirect(url_for("admin.manage_users"))


@admin_bp.route("/settings", methods=["GET", "POST"])
def settings():
    defaults = {
        "library_name": ("Al-Khwarizmi Smart Library", "Portal header title displayed to all users"),
        "hero_quote_text": ("Education is the most powerful weapon which you can use to change the world.", "Main quote displayed on the home page hero section"),
        "hero_quote_author": ("- Nelson Mandela", "Author of the home page hero quote"),
        "default_borrow_days": ("14", "Default duration (days) for a physical book borrow loan"),
        "max_active_borrows": ("5", "Maximum number of books a student can borrow at once"),
        "late_fine_per_day": ("1000", "Fine in UZS charged per day for overdue books"),
        "daily_request_limit": ("5", "Maximum number of borrow requests a user can submit per day"),
        "max_waiting_list_requests": ("3", "Maximum number of queued waiting-list requests per user"),
        "block_unpaid_fines": ("True", "Block new requests while unpaid fines exist (True/False)"),
        "allow_registration": ("True", "Allow public sign-ups on the registration page (True/False)"),
        "allow_digital_downloads": ("True", "Enable or disable digital PDF downloading globally (True/False)"),
        "homepage_announcements_enabled": ("True", "Enable or disable announcement cards on the home page (True/False)")
    }

    settings_list = Settings.query.all()
    existing_keys = {setting.key for setting in settings_list}
    created_defaults = False
    for key, (val, desc) in defaults.items():
        if key not in existing_keys:
            db.session.add(Settings(key=key, value=val, description=desc))
            created_defaults = True
    if created_defaults:
        db.session.commit()
        settings_list = Settings.query.all()

    settings_dict = {s.key: s for s in settings_list}

    banner = SiteBanner.query.first()
    if not banner:
        banner = SiteBanner(enabled=False, banner_text="", banner_type="info", banner_icon="📢", scroll_speed="normal")
        db.session.add(banner)
        db.session.commit()

    if request.method == "POST":
        for key in defaults.keys():
            form_val = request.form.get(key, "").strip()
            if key in settings_dict:
                settings_dict[key].value = form_val
                
        banner.enabled = request.form.get("banner_enabled") == "True"
        banner.banner_text = request.form.get("banner_text", "").strip()
        banner.banner_type = request.form.get("banner_type", "info")
        banner.banner_icon = request.form.get("banner_icon", "").strip() or None
        banner.scroll_speed = request.form.get("scroll_speed", "normal")
        
        db.session.commit()
        flash("System settings updated successfully.", "success")
        return redirect(url_for("admin.settings"))

    return render_template(
        "admin/settings.html",
        settings=settings_dict,
        banner=banner
    )


@admin_bp.route("/activity")
def activity():
    page = request.args.get("page", 1, type=int)
    q = request.args.get("q", "").strip()

    query = ActivityLog.query.outerjoin(User, ActivityLog.user_id == User.id)

    if q:
        query = query.filter(
            or_(
                ActivityLog.action.ilike(f"%{q}%"),
                User.fullname.ilike(f"%{q}%"),
                User.username.ilike(f"%{q}%"),
                User.email.ilike(f"%{q}%"),
                User.faculty.ilike(f"%{q}%"),
                User.group_name.ilike(f"%{q}%")
            )
        )

    pagination = query.order_by(
        desc(ActivityLog.created_at)
    ).paginate(
        page=page,
        per_page=20,
        error_out=False
    )

    log_count = query.count()

    return render_template(
        "admin/activity.html",
        pagination=pagination,
        q=q,
        log_count=log_count
    )


@admin_bp.route("/activities/export")
def export_activities():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash("Excel export requires openpyxl. Run pip install -r requirements.txt and try again.", "danger")
        return redirect(url_for("admin.activity"))

    q = request.args.get("q", "").strip()

    query = ActivityLog.query.outerjoin(User, ActivityLog.user_id == User.id)

    if q:
        query = query.filter(
            or_(
                ActivityLog.action.ilike(f"%{q}%"),
                User.fullname.ilike(f"%{q}%"),
                User.username.ilike(f"%{q}%"),
                User.email.ilike(f"%{q}%"),
                User.faculty.ilike(f"%{q}%"),
                User.group_name.ilike(f"%{q}%")
            )
        )

    logs = query.order_by(desc(ActivityLog.created_at)).all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Website Logs"

    headers = [
        "Log ID", "Timestamp", "Actor Type", "User ID", "Fullname",
        "Username", "Email", "Role", "Faculty", "Group", "Action"
    ]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="0B3B6F")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D9E2F1"),
        right=Side(style="thin", color="D9E2F1"),
        top=Side(style="thin", color="D9E2F1"),
        bottom=Side(style="thin", color="D9E2F1")
    )

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for log in logs:
        actor = log.user
        sheet.append([
            log.id,
            log.created_at.strftime("%Y-%m-%d %H:%M:%S") if log.created_at else "",
            "User" if actor else "System",
            actor.id if actor else "",
            actor.fullname if actor and actor.fullname else "System",
            actor.username if actor else "",
            actor.email if actor else "",
            actor.role if actor else "system",
            actor.faculty_display if actor else "",
            actor.group_name if actor and actor.group_name else "",
            log.action
        ])

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=11):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    widths = [10, 22, 14, 10, 26, 18, 30, 14, 22, 16, 46]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="website_logs.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@admin_bp.route("/faculties", methods=["GET", "POST"])
def manage_faculties():
    form = FacultyForm()
    edit_form = FacultyForm(prefix="edit")
    q = request.args.get("q", "").strip()

    if form.validate_on_submit() and not request.form.get("edit_id"):
        faculty_name = form.name.data.strip()
        existing = Faculty.query.filter(Faculty.name.ilike(faculty_name)).first()
        if existing:
            flash("Faculty already exists.", "warning")
        else:
            faculty = Faculty(name=faculty_name)
            db.session.add(faculty)
            db.session.commit()
            flash("Faculty added successfully.", "success")
        return redirect(url_for("admin.manage_faculties", q=q))

    faculties_query = Faculty.query
    if q:
        faculties_query = faculties_query.filter(Faculty.name.ilike(f"%{q}%"))

    faculties = faculties_query.order_by(Faculty.name.asc()).all()
    editing = None
    edit_id = request.args.get("edit", type=int)
    if edit_id:
        editing = Faculty.query.get_or_404(edit_id)
        edit_form.name.data = editing.name

    return render_template(
        "admin/faculties.html",
        faculties=faculties,
        form=form,
        edit_form=edit_form,
        editing=editing,
        q=q,
    )


@admin_bp.route("/faculties/<int:faculty_id>/edit", methods=["POST"])
def faculty_edit(faculty_id):
    q = request.args.get("q", "").strip()
    faculty = Faculty.query.get_or_404(faculty_id)
    form = FacultyForm(prefix="edit")

    if not form.validate_on_submit():
        flash("Could not update faculty. Check the form and try again.", "danger")
        return redirect(url_for("admin.manage_faculties", q=q, edit=faculty_id))

    new_name = form.name.data.strip()
    duplicate = Faculty.query.filter(
        Faculty.id != faculty.id,
        Faculty.name.ilike(new_name),
    ).first()
    if duplicate:
        flash("Another faculty with this name already exists.", "warning")
        return redirect(url_for("admin.manage_faculties", q=q, edit=faculty_id))

    faculty.name = new_name
    User.query.filter_by(faculty_id=faculty.id).update(
        {User.faculty: new_name},
        synchronize_session=False,
    )
    db.session.commit()
    flash("Faculty updated successfully.", "success")
    return redirect(url_for("admin.manage_faculties", q=q))


@admin_bp.route("/faculties/<int:faculty_id>/delete", methods=["POST"])
def faculty_delete(faculty_id):
    q = request.args.get("q", "").strip()
    faculty = Faculty.query.get_or_404(faculty_id)
    users_count = User.query.filter_by(faculty_id=faculty.id).count()

    if users_count:
        flash("Cannot delete faculty while users are assigned to it.", "danger")
    else:
        db.session.delete(faculty)
        db.session.commit()
        flash("Faculty deleted successfully.", "success")

    return redirect(url_for("admin.manage_faculties", q=q))


@admin_bp.route("/announcements")
def manage_announcements():
    from sqlalchemy import case
    priority_order = case(
        (Announcement.priority == "high", 1),
        (Announcement.priority == "medium", 2),
        (Announcement.priority == "low", 3),
        else_=4
    )
    announcements = Announcement.query.order_by(priority_order, Announcement.created_at.desc()).all()
    return render_template("admin/manage_announcements.html", announcements=announcements)


@admin_bp.route("/announcements/create", methods=["GET", "POST"])
def create_announcement():
    import os
    from app.routes.librarian import upload_file
    
    if request.method == "POST":
        title = request.form.get("title", "").strip()
        description = request.form.get("description", "").strip()
        link = request.form.get("link", "").strip() or None
        priority = request.form.get("priority", "medium")
        status = request.form.get("status", "active")
        
        start_date_str = request.form.get("start_date", "").strip()
        end_date_str = request.form.get("end_date", "").strip()
        
        start_date = None
        if start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
            except ValueError:
                pass
                
        end_date = None
        if end_date_str:
            try:
                end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
            except ValueError:
                pass
                
        image_file = request.files.get("image")
        image_path = None
        if image_file and image_file.filename:
            image_path = upload_file(image_file, "announcements")
            
        if not title or not description:
            flash("Title and description are required.", "danger")
            return redirect(url_for("admin.create_announcement"))
            
        announcement = Announcement(
            title=title,
            description=description,
            image=image_path,
            link=link,
            priority=priority,
            status=status,
            start_date=start_date,
            end_date=end_date
        )
        db.session.add(announcement)
        db.session.commit()
        flash("Announcement created successfully.", "success")
        return redirect(url_for("admin.manage_announcements"))
        
    return render_template("admin/announcement_form.html", announcement=None)


@admin_bp.route("/announcements/<int:id>/edit", methods=["GET", "POST"])
def edit_announcement(id):
    import os
    from app.routes.librarian import upload_file
    announcement = Announcement.query.get_or_404(id)
    
    if request.method == "POST":
        title = request.form.get("title", "").strip()
        description = request.form.get("description", "").strip()
        link = request.form.get("link", "").strip() or None
        priority = request.form.get("priority", "medium")
        status = request.form.get("status", "active")
        
        start_date_str = request.form.get("start_date", "").strip()
        end_date_str = request.form.get("end_date", "").strip()
        
        start_date = None
        if start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
            except ValueError:
                pass
                
        end_date = None
        if end_date_str:
            try:
                end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
            except ValueError:
                pass
                
        if not title or not description:
            flash("Title and description are required.", "danger")
            return redirect(url_for("admin.edit_announcement", id=id))
            
        image_file = request.files.get("image")
        if image_file and image_file.filename:
            if announcement.image:
                old_image_path = os.path.join(current_app.root_path, "static", announcement.image)
                if os.path.exists(old_image_path):
                    try:
                        os.remove(old_image_path)
                    except Exception:
                        pass
            announcement.image = upload_file(image_file, "announcements")
        elif request.form.get("remove_image") == "true":
            if announcement.image:
                old_image_path = os.path.join(current_app.root_path, "static", announcement.image)
                if os.path.exists(old_image_path):
                    try:
                        os.remove(old_image_path)
                    except Exception:
                        pass
            announcement.image = None
            
        announcement.title = title
        announcement.description = description
        announcement.link = link
        announcement.priority = priority
        announcement.status = status
        announcement.start_date = start_date
        announcement.end_date = end_date
        
        db.session.commit()
        flash("Announcement updated successfully.", "success")
        return redirect(url_for("admin.manage_announcements"))
        
    return render_template("admin/announcement_form.html", announcement=announcement)


@admin_bp.route("/announcements/<int:id>/delete", methods=["POST"])
def delete_announcement(id):
    import os
    announcement = Announcement.query.get_or_404(id)
    if announcement.image:
        image_path = os.path.join(current_app.root_path, "static", announcement.image)
        if os.path.exists(image_path):
            try:
                os.remove(image_path)
            except Exception:
                pass
    db.session.delete(announcement)
    db.session.commit()
    flash("Announcement deleted successfully.", "success")
    return redirect(url_for("admin.manage_announcements"))
