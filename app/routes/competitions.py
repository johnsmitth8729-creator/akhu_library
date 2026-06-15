from __future__ import annotations

import os
import uuid
import csv
from io import BytesIO
from io import StringIO

from flask import Blueprint, current_app, jsonify, redirect, render_template, request, send_file, url_for, flash
from flask_login import current_user, login_required
from sqlalchemy import func

from app import db
from app.forms.competition_forms import CompetitionForm, parse_datetime_local
from app.models import (
    CompetitionBook,
    CompetitionCertificate,
    CompetitionFaculty,
    CompetitionGroup,
    CompetitionQuestion,
    Faculty,
    Question,
    QuestionOption,
    QuizAttempt,
    QuizViolation,
    ReadingCompetition,
)
from app.services import competition_service
from app.services.ranking_service import get_leaderboard, get_user_competition_stats
from app.utils.datetime import now_local


competitions_bp = Blueprint("competitions", __name__)


def staff_required():
    if not current_user.is_authenticated or not current_user.is_staff:
        flash("Staff access required.", "danger")
        return redirect(url_for("main.home"))
    return None


def _populate_form_choices(form: CompetitionForm) -> None:
    form.book_ids.choices = []
    form.faculty_ids.choices = [
        (faculty.id, faculty.name)
        for faculty in Faculty.query.order_by(Faculty.name.asc()).all()
    ]
    form.question_ids.choices = []


def _set_form_defaults(form: CompetitionForm, competition: ReadingCompetition) -> None:
    form.title.data = competition.title
    form.description.data = competition.description
    form.competition_type.data = competition.competition_type
    form.visibility.data = competition.visibility
    form.start_date.data = competition.start_date.strftime("%Y-%m-%dT%H:%M")
    form.end_date.data = competition.end_date.strftime("%Y-%m-%dT%H:%M")
    form.max_attempts.data = competition.max_attempts
    form.passing_score.data = competition.passing_score
    form.top_winners_count.data = competition.top_winners_count
    form.time_limit_minutes.data = competition.time_limit_minutes
    form.randomize_questions.data = competition.randomize_questions
    form.randomize_answers.data = competition.randomize_answers
    form.prevent_reopen_completed.data = competition.prevent_reopen_completed
    form.secure_quiz_mode.data = competition.secure_quiz_mode
    form.enable_watermark.data = competition.enable_watermark
    form.disable_copy.data = competition.disable_copy
    form.disable_print.data = competition.disable_print
    form.require_fullscreen.data = competition.require_fullscreen
    form.track_focus_loss.data = competition.track_focus_loss
    form.track_devtools.data = competition.track_devtools
    form.book_ids.data = [row.book_id for row in competition.books]
    form.faculty_ids.data = [row.faculty_id for row in competition.faculty_restrictions]
    form.group_names.data = ", ".join(row.group_name for row in competition.group_restrictions)
    form.question_ids.data = [row.question_id for row in competition.competition_questions]


def _save_competition_image(file_storage) -> str | None:
    if not file_storage or not file_storage.filename:
        return None
    ext = os.path.splitext(file_storage.filename)[1].lower()
    if ext not in {".jpg", ".jpeg", ".png", ".webp"}:
        raise ValueError("Competition image must be JPG, PNG, or WEBP.")

    rel_dir = os.path.join("uploads", "competitions")
    abs_dir = os.path.join(current_app.static_folder, rel_dir)
    os.makedirs(abs_dir, exist_ok=True)
    filename = f"competition_{uuid.uuid4().hex}{ext}"
    file_storage.save(os.path.join(abs_dir, filename))
    return f"{rel_dir}/{filename}".replace("\\", "/")


def _normalise_correct_answers(raw: str, question_type: str) -> set[str]:
    value = (raw or "").strip().lower()
    if question_type == Question.TYPE_TRUE_FALSE:
        if value in {"true", "yes", "a", "1"}:
            return {"A"}
        if value in {"false", "no", "b", "2"}:
            return {"B"}
        raise ValueError("True/False correct answer must be True or False.")

    letters = {item.strip().upper() for item in value.replace(";", ",").split(",") if item.strip()}
    aliases = {"1": "A", "2": "B", "3": "C", "4": "D"}
    letters = {aliases.get(letter, letter) for letter in letters}
    if not letters or not letters.issubset({"A", "B", "C", "D"}):
        raise ValueError("Correct answer must use A, B, C, D letters.")
    if question_type == Question.TYPE_SINGLE and len(letters) != 1:
        raise ValueError("Single choice questions must have exactly one correct answer.")
    return letters


def _parse_question_rows_from_request() -> tuple[list[dict], list[str]]:
    texts = request.form.getlist("question_text")
    types = request.form.getlist("question_type")
    option_a = request.form.getlist("option_a")
    option_b = request.form.getlist("option_b")
    option_c = request.form.getlist("option_c")
    option_d = request.form.getlist("option_d")
    correct_answers = request.form.getlist("correct_answer")
    points = request.form.getlist("points")

    rows = []
    errors = []
    seen = set()

    for index, text in enumerate(texts):
        question_text = (text or "").strip()
        if not question_text:
            continue
        lowered = question_text.lower()
        if lowered in seen:
            errors.append(f"Question {index + 1}: duplicate question skipped.")
            continue
        seen.add(lowered)

        question_type = types[index] if index < len(types) else Question.TYPE_SINGLE
        if question_type not in {Question.TYPE_SINGLE, Question.TYPE_MULTIPLE_CHOICE, Question.TYPE_TRUE_FALSE}:
            errors.append(f"Question {index + 1}: invalid question type.")
            continue

        options = [
            option_a[index] if index < len(option_a) else "",
            option_b[index] if index < len(option_b) else "",
            option_c[index] if index < len(option_c) else "",
            option_d[index] if index < len(option_d) else "",
        ]
        if question_type == Question.TYPE_TRUE_FALSE:
            options = ["True", "False", "", ""]

        clean_options = [value.strip() for value in options]
        required_options = clean_options[:2] if question_type == Question.TYPE_TRUE_FALSE else clean_options
        if sum(1 for value in required_options if value) < 2:
            errors.append(f"Question {index + 1}: at least two options are required.")
            continue

        try:
            correct = _normalise_correct_answers(
                correct_answers[index] if index < len(correct_answers) else "",
                question_type,
            )
            point_value = int(points[index] if index < len(points) and points[index] else 1)
        except ValueError as exc:
            errors.append(f"Question {index + 1}: {exc}")
            continue

        if point_value < 1:
            errors.append(f"Question {index + 1}: points must be at least 1.")
            continue

        rows.append(
            {
                "question_text": question_text,
                "question_type": question_type,
                "options": clean_options,
                "correct": correct,
                "points": point_value,
            }
        )

    return rows, errors


def _parse_question_rows_from_excel(file_storage) -> tuple[list[dict], dict]:
    report = {"imported": 0, "skipped": 0, "invalid": 0, "errors": []}
    if not file_storage or not file_storage.filename:
        return [], report

    try:
        import openpyxl
        workbook = openpyxl.load_workbook(file_storage, read_only=True, data_only=True)
    except Exception as exc:
        report["invalid"] += 1
        report["errors"].append(f"Could not read Excel file: {exc}")
        return [], report

    sheet = workbook.active
    rows = []
    seen = set()
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        values = [str(cell).strip() if cell is not None else "" for cell in row[:7]]
        if row_number == 1 and values and values[0].lower() in {"question", "question text", "question_text"}:
            continue
        if not any(values):
            report["skipped"] += 1
            continue

        question_text = values[0] if len(values) > 0 else ""
        if not question_text:
            report["skipped"] += 1
            continue
        lowered = question_text.lower()
        if lowered in seen:
            report["skipped"] += 1
            report["errors"].append(f"Row {row_number}: duplicate question skipped.")
            continue
        seen.add(lowered)

        options = [
            values[1] if len(values) > 1 else "",
            values[2] if len(values) > 2 else "",
            values[3] if len(values) > 3 else "",
            values[4] if len(values) > 4 else "",
        ]
        correct_raw = values[5] if len(values) > 5 else ""
        points_raw = values[6] if len(values) > 6 else "1"
        question_type = Question.TYPE_MULTIPLE_CHOICE if "," in correct_raw else Question.TYPE_SINGLE

        try:
            correct = _normalise_correct_answers(correct_raw, question_type)
            point_value = int(points_raw or 1)
        except ValueError as exc:
            report["invalid"] += 1
            report["errors"].append(f"Row {row_number}: {exc}")
            continue

        if sum(1 for value in options if value) < 2:
            report["invalid"] += 1
            report["errors"].append(f"Row {row_number}: at least two options required.")
            continue

        rows.append(
            {
                "question_text": question_text,
                "question_type": question_type,
                "options": options,
                "correct": correct,
                "points": max(point_value, 1),
            }
        )
        report["imported"] += 1

    return rows, report


def _replace_competition_questions(competition: ReadingCompetition, rows: list[dict]) -> None:
    old_question_ids = [link.question_id for link in competition.competition_questions]
    CompetitionQuestion.query.filter_by(competition_id=competition.id).delete()
    db.session.flush()

    for question_id in old_question_ids:
        if not QuizAttempt.query.filter_by(competition_id=competition.id).first():
            question = Question.query.get(question_id)
            if question:
                db.session.delete(question)

    for index, row in enumerate(rows, start=1):
        question = Question(
            question_text=row["question_text"],
            question_type=row["question_type"],
            points=row["points"],
            difficulty="medium",
            is_active=True,
            is_archived=False,
            created_by_id=current_user.id,
        )
        db.session.add(question)
        db.session.flush()

        for option_index, option_text in enumerate(row["options"][:4]):
            if not option_text:
                continue
            letter = "ABCD"[option_index]
            db.session.add(
                QuestionOption(
                    question_id=question.id,
                    option_text=option_text,
                    is_correct=letter in row["correct"],
                    sort_order=option_index,
                )
            )

        db.session.add(
            CompetitionQuestion(
                competition_id=competition.id,
                question_id=question.id,
                sort_order=index,
            )
        )


def _competition_questions_payload(competition: ReadingCompetition | None) -> list[dict]:
    if not competition:
        return []
    payload = []
    for link in competition.competition_questions:
        question = link.question
        if not question:
            continue
        options = list(question.options)
        option_values = ["", "", "", ""]
        correct = []
        for index, option in enumerate(options[:4]):
            option_values[index] = option.option_text
            if option.is_correct:
                correct.append("ABCD"[index])
        payload.append(
            {
                "question_text": question.question_text,
                "question_type": question.question_type,
                "option_a": option_values[0],
                "option_b": option_values[1],
                "option_c": option_values[2],
                "option_d": option_values[3],
                "correct_answer": ",".join(correct),
                "points": question.points,
            }
        )
    return payload


def _save_competition_from_form(
    form: CompetitionForm,
    competition: ReadingCompetition | None = None,
) -> ReadingCompetition:
    start_date = parse_datetime_local(form.start_date.data)
    end_date = parse_datetime_local(form.end_date.data)
    if start_date >= end_date:
        raise ValueError("Start date must be before end date.")

    if competition is None:
        competition = ReadingCompetition(created_by_id=current_user.id)
        db.session.add(competition)

    competition.title = form.title.data.strip()
    competition.description = form.description.data.strip() if form.description.data else None
    competition.competition_type = form.competition_type.data
    competition.visibility = form.visibility.data
    competition.start_date = start_date
    competition.end_date = end_date
    competition.max_attempts = form.max_attempts.data or 1
    competition.passing_score = form.passing_score.data or 60
    competition.top_winners_count = form.top_winners_count.data or 3
    competition.time_limit_minutes = form.time_limit_minutes.data or None
    competition.randomize_questions = bool(form.randomize_questions.data)
    competition.randomize_answers = bool(form.randomize_answers.data)
    competition.prevent_reopen_completed = bool(form.prevent_reopen_completed.data)
    competition.secure_quiz_mode = bool(form.secure_quiz_mode.data)
    competition.enable_watermark = bool(form.enable_watermark.data)
    competition.disable_copy = bool(form.disable_copy.data)
    competition.disable_print = bool(form.disable_print.data)
    competition.require_fullscreen = bool(form.require_fullscreen.data)
    competition.track_focus_loss = bool(form.track_focus_loss.data)
    competition.track_devtools = bool(form.track_devtools.data)
    competition.status = request.form.get("status") or competition.status or ReadingCompetition.STATUS_DRAFT

    image_path = _save_competition_image(form.competition_image.data)
    if image_path:
        competition.image_path = image_path

    db.session.flush()

    CompetitionBook.query.filter_by(competition_id=competition.id).delete()
    CompetitionFaculty.query.filter_by(competition_id=competition.id).delete()
    CompetitionGroup.query.filter_by(competition_id=competition.id).delete()
    db.session.flush()

    for book_id in dict.fromkeys(form.book_ids.data or []):
        db.session.add(CompetitionBook(competition_id=competition.id, book_id=book_id))

    for faculty_id in dict.fromkeys(form.faculty_ids.data or []):
        db.session.add(CompetitionFaculty(competition_id=competition.id, faculty_id=faculty_id))

    if competition.competition_type == ReadingCompetition.TYPE_GROUP:
        groups = [
            value.strip()
            for value in (form.group_names.data or "").split(",")
            if value.strip()
        ]
        for group_name in dict.fromkeys(groups):
            db.session.add(CompetitionGroup(competition_id=competition.id, group_name=group_name))

    question_rows, form_errors = _parse_question_rows_from_request()
    excel_rows, import_report = _parse_question_rows_from_excel(request.files.get("questions_excel"))
    question_rows.extend(excel_rows)

    if import_report["imported"] or import_report["skipped"] or import_report["invalid"]:
        flash(
            f"Excel import: {import_report['imported']} imported, "
            f"{import_report['skipped']} skipped, {import_report['invalid']} invalid.",
            "info",
        )
    for error in form_errors + import_report["errors"][:5]:
        flash(error, "warning")

    has_attempts = QuizAttempt.query.filter_by(competition_id=competition.id).count() > 0
    if question_rows and not has_attempts:
        _replace_competition_questions(competition, question_rows)
    elif question_rows and has_attempts:
        flash("Questions were not changed because this competition already has attempts.", "warning")
    elif not competition.competition_questions:
        raise ValueError("Add at least one valid question.")

    db.session.commit()
    return competition


@competitions_bp.route("/")
@login_required
def index():
    competitions = (
        ReadingCompetition.query.filter(
            ReadingCompetition.status.in_(
                [ReadingCompetition.STATUS_PUBLISHED, ReadingCompetition.STATUS_CLOSED]
            )
        )
        .filter(ReadingCompetition.status != ReadingCompetition.STATUS_ARCHIVED)
        .order_by(ReadingCompetition.start_date.desc())
        .all()
    )
    items = [
        {
            "competition": competition,
            "best": competition_service.get_best_completed_attempt(current_user.id, competition.id),
        }
        for competition in competitions
    ]
    stats = get_user_competition_stats(current_user.id)
    return render_template("competitions/index.html", items=items, stats=stats, utcnow=now_local)


@competitions_bp.route("/<int:competition_id>")
@login_required
def detail(competition_id: int):
    competition = ReadingCompetition.query.get_or_404(competition_id)
    can_enter, message = competition_service.user_can_enter_competition(current_user, competition)
    can_start, start_msg = competition_service.can_start_attempt(current_user, competition)
    leaderboard = get_leaderboard(competition.id, limit=5)
    attempts_used = competition_service.get_user_attempt_count(current_user.id, competition.id)
    my_best = competition_service.get_best_completed_attempt(current_user.id, competition.id)
    return render_template(
        "competitions/detail.html",
        competition=competition,
        can_enter=can_enter,
        can_start=can_start,
        message=message,
        start_msg=start_msg,
        leaderboard=leaderboard,
        attempts_used=attempts_used,
        my_best=my_best,
    )


@competitions_bp.route("/<int:competition_id>/start", methods=["POST"])
@login_required
def start_quiz(competition_id: int):
    competition = ReadingCompetition.query.get_or_404(competition_id)
    try:
        attempt = competition_service.start_quiz_attempt(current_user, competition)
    except ValueError as exc:
        flash(str(exc), "warning")
        return redirect(url_for("competitions.detail", competition_id=competition.id))
    return redirect(url_for("competitions.quiz", attempt_id=attempt.id))


@competitions_bp.route("/attempt/<int:attempt_id>")
@login_required
def quiz(attempt_id: int):
    attempt = QuizAttempt.query.get_or_404(attempt_id)
    if attempt.user_id != current_user.id and not current_user.is_staff:
        flash("You cannot open this attempt.", "danger")
        return redirect(url_for("competitions.index"))
    if attempt.status != QuizAttempt.STATUS_IN_PROGRESS:
        return redirect(url_for("competitions.result", attempt_id=attempt.id))

    questions = []
    for question in competition_service.get_attempt_questions(attempt):
        option_ids = competition_service.build_options_order(
            question,
            attempt.competition.randomize_answers,
        )
        options_by_id = {option.id: option for option in question.options}
        questions.append(
            {
                "question": question,
                "options": [options_by_id[option_id] for option_id in option_ids if option_id in options_by_id],
            }
        )

    watermark = {
        "name": current_user.fullname or current_user.username,
        "username": current_user.username,
        "group": current_user.group_name or "",
        "time": now_local().strftime("%Y-%m-%d %H:%M"),
    }
    return render_template(
        "competitions/quiz.html",
        attempt=attempt,
        competition=attempt.competition,
        questions=questions,
        watermark=watermark,
    )


@competitions_bp.route("/attempt/<int:attempt_id>/submit", methods=["POST"])
@login_required
def submit_quiz(attempt_id: int):
    attempt = QuizAttempt.query.get_or_404(attempt_id)
    if attempt.user_id != current_user.id:
        flash("You cannot submit this attempt.", "danger")
        return redirect(url_for("competitions.index"))

    payload = {}
    for key, values in request.form.lists():
        if key.startswith("q_"):
            payload[key[2:]] = [int(value) for value in values if str(value).isdigit()]

    try:
        competition_service.submit_attempt(attempt, payload)
    except ValueError as exc:
        flash(str(exc), "warning")
    return redirect(url_for("competitions.result", attempt_id=attempt.id))


@competitions_bp.route("/attempt/<int:attempt_id>/autosave", methods=["POST"])
@login_required
def autosave_quiz(attempt_id: int):
    attempt = QuizAttempt.query.get_or_404(attempt_id)
    if attempt.user_id != current_user.id:
        return jsonify({"ok": False}), 403

    data = request.get_json(silent=True) or {}
    saved = 0
    question_ids = [question.id for question in competition_service.get_attempt_questions(attempt)]
    for question_id in question_ids:
        values = data.get(f"q_{question_id}", [])
        try:
            option_ids = [int(value) for value in values]
        except (TypeError, ValueError):
            option_ids = []
        if competition_service.save_draft_answer(attempt.id, question_id, option_ids):
            saved += 1
    return jsonify({"ok": True, "saved": saved})


@competitions_bp.route("/attempt/<int:attempt_id>/violation", methods=["POST"])
@login_required
def report_violation(attempt_id: int):
    attempt = QuizAttempt.query.get_or_404(attempt_id)
    if attempt.user_id != current_user.id:
        return jsonify({"ok": False}), 403

    data = request.get_json(silent=True) or {}
    event_type = (data.get("type") or "unknown")[:50]
    details = data.get("details")
    db.session.add(QuizViolation(attempt_id=attempt.id, event_type=event_type, details=details))
    attempt.violation_count = (attempt.violation_count or 0) + 1
    if event_type == "focus_loss":
        attempt.focus_loss_count = (attempt.focus_loss_count or 0) + 1
    if event_type == "fullscreen_exit":
        attempt.fullscreen_exit_count = (attempt.fullscreen_exit_count or 0) + 1
    db.session.commit()
    return jsonify({"ok": True})


@competitions_bp.route("/attempt/<int:attempt_id>/result")
@login_required
def result(attempt_id: int):
    attempt = QuizAttempt.query.get_or_404(attempt_id)
    if attempt.user_id != current_user.id and not current_user.is_staff:
        flash("You cannot view this result.", "danger")
        return redirect(url_for("competitions.index"))

    review_items = []
    if current_user.is_staff:
        for answer in attempt.answers:
            correct_options = [opt.option_text for opt in answer.question.options if opt.is_correct]
            review_items.append(
                {
                    "question_text": answer.question.question_text,
                    "correct": answer.is_correct,
                    "correct_answer": ", ".join(correct_options),
                }
            )
    return render_template(
        "competitions/result.html",
        attempt=attempt,
        competition=attempt.competition,
        passed=attempt.passed,
        cert=attempt.certificate,
        review_items=review_items,
    )


@competitions_bp.route("/<int:competition_id>/leaderboard")
@login_required
def leaderboard(competition_id: int):
    competition = ReadingCompetition.query.get_or_404(competition_id)
    rows = get_leaderboard(competition.id, limit=100)
    my_rank = competition_service.get_best_completed_attempt(current_user.id, competition.id)
    return render_template(
        "competitions/leaderboard.html",
        competition=competition,
        rows=rows,
        my_rank=my_rank,
    )


@competitions_bp.route("/certificate/<int:cert_id>/download")
@login_required
def download_certificate(cert_id: int):
    cert = CompetitionCertificate.query.get_or_404(cert_id)
    if cert.user_id != current_user.id and not current_user.is_staff:
        flash("You cannot download this certificate.", "danger")
        return redirect(url_for("competitions.index"))
    if cert.pdf_path:
        path = cert.pdf_path
        if not os.path.isabs(path):
            path = os.path.join(current_app.static_folder, path)
        return send_file(path, as_attachment=True)
    flash("Certificate file is not ready yet.", "warning")
    return redirect(url_for("competitions.result", attempt_id=cert.attempt_id))


@competitions_bp.route("/verify/<string:code>")
def verify_certificate(code: str):
    cert = CompetitionCertificate.query.filter_by(
        verification_code=(code or "").strip().upper(),
    ).first_or_404()
    return render_template("competitions/verify.html", cert=cert)


@competitions_bp.route("/manage")
@login_required
def manage_list():
    err = staff_required()
    if err:
        return err
    competitions = (
        ReadingCompetition.query.filter(ReadingCompetition.status != ReadingCompetition.STATUS_ARCHIVED)
        .order_by(ReadingCompetition.created_at.desc())
        .all()
    )
    return render_template("competitions/manage/list.html", competitions=competitions)


@competitions_bp.route("/manage/create", methods=["GET", "POST"])
@login_required
def manage_create():
    err = staff_required()
    if err:
        return err
    return _manage_form(None)


@competitions_bp.route("/manage/<int:competition_id>/edit", methods=["GET", "POST"])
@login_required
def manage_edit(competition_id: int):
    err = staff_required()
    if err:
        return err
    competition = ReadingCompetition.query.get_or_404(competition_id)
    return _manage_form(competition)


def _manage_form(competition: ReadingCompetition | None):
    form = CompetitionForm()
    _populate_form_choices(form)

    if request.method == "GET" and competition:
        _set_form_defaults(form, competition)

    form.book_ids.data = form.book_ids.data or []
    form.faculty_ids.data = form.faculty_ids.data or []
    form.question_ids.data = form.question_ids.data or []

    if form.validate_on_submit():
        try:
            competition = _save_competition_from_form(form, competition)
        except ValueError as exc:
            flash(str(exc), "danger")
        else:
            flash("Competition saved.", "success")
            return redirect(url_for("competitions.manage_list"))

    question_payload = _competition_questions_payload(competition)
    return render_template(
        "competitions/manage/form.html",
        form=form,
        competition=competition,
        question_payload=question_payload,
        has_attempts=bool(competition and QuizAttempt.query.filter_by(competition_id=competition.id).count()),
    )


@competitions_bp.route("/manage/<int:competition_id>/publish", methods=["POST"])
@login_required
def publish_competition(competition_id: int):
    err = staff_required()
    if err:
        return err
    competition = ReadingCompetition.query.get_or_404(competition_id)
    if competition.start_date >= competition.end_date:
        flash("Start date must be before end date.", "danger")
    elif not competition.competition_questions:
        flash("Add at least one question before publishing.", "warning")
    else:
        competition.status = ReadingCompetition.STATUS_PUBLISHED
        db.session.commit()
        flash("Competition published.", "success")
    return redirect(url_for("competitions.manage_list"))


@competitions_bp.route("/manage/<int:competition_id>/close", methods=["POST"])
@login_required
def close_competition(competition_id: int):
    err = staff_required()
    if err:
        return err
    competition = ReadingCompetition.query.get_or_404(competition_id)
    competition.status = ReadingCompetition.STATUS_CLOSED
    db.session.commit()
    flash("Competition closed.", "success")
    return redirect(url_for("competitions.manage_list"))


@competitions_bp.route("/manage/<int:competition_id>/results")
@login_required
def manage_results(competition_id: int):
    err = staff_required()
    if err:
        return err
    competition = ReadingCompetition.query.get_or_404(competition_id)
    attempts = (
        QuizAttempt.query.filter_by(competition_id=competition.id)
        .order_by(
            QuizAttempt.rank_position.is_(None),
            QuizAttempt.rank_position.asc(),
            QuizAttempt.started_at.desc(),
        )
        .all()
    )
    violations = (
        QuizViolation.query.join(QuizAttempt)
        .filter(QuizAttempt.competition_id == competition.id)
        .order_by(QuizViolation.created_at.desc())
        .all()
    )
    return render_template(
        "competitions/manage/results.html",
        competition=competition,
        attempts=attempts,
        violations=violations,
    )


@competitions_bp.route("/manage/<int:competition_id>/export")
@login_required
def export_results(competition_id: int):
    err = staff_required()
    if err:
        return err

    competition = ReadingCompetition.query.get_or_404(competition_id)
    attempts = QuizAttempt.query.filter_by(competition_id=competition.id).all()
    columns = ["Name", "Username", "Faculty", "Group", "Competition", "Score", "Percentage", "Rank", "Started At", "Finished At"]
    rows = [
        [
            attempt.user.fullname if attempt.user else "",
            attempt.user.username if attempt.user else "",
            attempt.user.faculty_display if attempt.user else "",
            attempt.user.group_name if attempt.user else "",
            competition.title,
            attempt.score,
            attempt.percentage,
            attempt.rank_position,
            attempt.started_at.strftime("%Y-%m-%d %H:%M") if attempt.started_at else "",
            attempt.completed_at.strftime("%Y-%m-%d %H:%M") if attempt.completed_at else "",
        ]
        for attempt in attempts
    ]

    if request.args.get("format") == "csv":
        buffer = StringIO()
        writer = csv.writer(buffer)
        writer.writerow(columns)
        writer.writerows(rows)
        return send_file(
            BytesIO(buffer.getvalue().encode("utf-8-sig")),
            as_attachment=True,
            download_name=f"competition_{competition.id}_results.csv",
            mimetype="text/csv",
        )

    try:
        import openpyxl
    except ImportError:
        flash("openpyxl is required for export.", "danger")
        return redirect(url_for("competitions.manage_results", competition_id=competition.id))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append(columns)
    for row in rows:
        ws.append(row)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"competition_{competition.id}_results.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@competitions_bp.route("/manage/questions")
@login_required
def manage_questions():
    return redirect(url_for("questions.index"))


@competitions_bp.route("/analytics")
@login_required
def analytics():
    err = staff_required()
    if err:
        return err

    total = ReadingCompetition.query.count()
    published = ReadingCompetition.query.filter_by(status=ReadingCompetition.STATUS_PUBLISHED).count()
    participants = db.session.query(func.count(func.distinct(QuizAttempt.user_id))).scalar() or 0
    completed = QuizAttempt.query.filter_by(status=QuizAttempt.STATUS_COMPLETED).all()
    average_score = round(sum(a.percentage for a in completed) / len(completed), 1) if completed else 0

    popular = (
        db.session.query(ReadingCompetition.title, func.count(QuizAttempt.id))
        .outerjoin(QuizAttempt, QuizAttempt.competition_id == ReadingCompetition.id)
        .group_by(ReadingCompetition.id)
        .order_by(func.count(QuizAttempt.id).desc())
        .limit(5)
        .all()
    )

    from app.models import User

    top_participants = (
        db.session.query(User.fullname, func.count(QuizAttempt.id), func.max(QuizAttempt.percentage))
        .join(QuizAttempt, QuizAttempt.user_id == User.id)
        .group_by(User.id)
        .order_by(func.count(QuizAttempt.id).desc())
        .limit(5)
        .all()
    )

    distribution = {"0_20": 0, "21_40": 0, "41_60": 0, "61_80": 0, "81_100": 0}
    for attempt in completed:
        pct = attempt.percentage or 0
        if pct <= 20:
            distribution["0_20"] += 1
        elif pct <= 40:
            distribution["21_40"] += 1
        elif pct <= 60:
            distribution["41_60"] += 1
        elif pct <= 80:
            distribution["61_80"] += 1
        else:
            distribution["81_100"] += 1

    return render_template(
        "competitions/analytics.html",
        data={
            "total_competitions": total,
            "published_competitions": published,
            "participants": participants,
            "average_score": average_score,
            "popular_competitions": popular,
            "top_participants": top_participants,
            "score_distribution": distribution,
        },
    )
