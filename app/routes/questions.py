import os
import uuid
from io import BytesIO
from flask import Blueprint, render_template, redirect, url_for, flash, request, current_app, jsonify, send_file, session
from flask_login import login_required, current_user
from sqlalchemy import desc

from app import db
from app.models import Question, QuestionCategory, QuestionOption, Book
from app.forms.competition_forms import QuestionForm
from app.services import question_service

questions_bp = Blueprint("questions", __name__)


def staff_required():
    if not current_user.is_authenticated or not current_user.is_staff:
        flash("Staff access required.", "danger")
        return redirect(url_for("main.home"))
    return None


@questions_bp.route("/")
@login_required
def index():
    err = staff_required()
    if err:
        return err

    q = request.args.get("q", "").strip()
    category_id = request.args.get("category_id", type=int)
    qtype = request.args.get("type", "").strip()
    difficulty = request.args.get("difficulty", "").strip()

    query = Question.query.filter_by(is_archived=False)
    if q:
        query = query.filter(Question.question_text.ilike(f"%{q}%"))
    if category_id:
        query = query.filter(Question.category_id == category_id)
    if qtype:
        query = query.filter(Question.question_type == qtype)
    if difficulty:
        query = query.filter(Question.difficulty == difficulty)

    questions = query.order_by(desc(Question.created_at)).all()
    total = Question.query.filter_by(is_archived=False).count()
    categories = QuestionCategory.query.filter_by(is_active=True).all()
    books = Book.query.order_by(Book.title.asc()).all()

    return render_template(
        "competitions/manage/questions.html",
        questions=questions,
        total=total,
        categories=categories,
        books=books,
        q=q,
        category_id=category_id,
        qtype=qtype,
        difficulty=difficulty,
    )


@questions_bp.route("/create", methods=["GET", "POST"])
@login_required
def create():
    err = staff_required()
    if err:
        return err

    form = QuestionForm()
    form.category_id.choices = [(0, "— None —")] + [
        (c.id, c.name) for c in QuestionCategory.query.filter_by(is_active=True).order_by(QuestionCategory.name.asc()).all()
    ]
    form.book_id.choices = [(0, "— None —")] + [
        (b.id, b.title) for b in Book.query.order_by(Book.title.asc()).all()
    ]

    if form.validate_on_submit():
        # Handle file upload
        image_path = None
        if form.image_file.data:
            file = form.image_file.data
            ext = os.path.splitext(file.filename)[1]
            safe_filename = f"{uuid.uuid4().hex}{ext}"
            rel_dir = os.path.join("uploads", "questions")
            abs_dir = os.path.join(current_app.static_folder, rel_dir)
            os.makedirs(abs_dir, exist_ok=True)
            file.save(os.path.join(abs_dir, safe_filename))
            image_path = f"{rel_dir}/{safe_filename}"

        # Parse options
        options_data = []
        if form.question_type.data == Question.TYPE_TRUE_FALSE:
            correct_val = request.form.get("correct_option_tf", "True")
            options_data = [
                {"option_text": "True", "is_correct": (correct_val == "True")},
                {"option_text": "False", "is_correct": (correct_val == "False")},
            ]
        else:
            options = request.form.getlist("option_text")
            correct_idxs = {int(x) for x in request.form.getlist("correct_option")}
            for idx, text in enumerate(options):
                if text.strip():
                    options_data.append({
                        "option_text": text.strip(),
                        "is_correct": idx in correct_idxs
                    })

        q = question_service.create_question(
            question_text=form.question_text.data,
            question_type=form.question_type.data,
            category_id=form.category_id.data,
            points=form.points.data,
            explanation=form.explanation.data,
            book_id=form.book_id.data,
            difficulty=form.difficulty.data,
            options_data=options_data,
            created_by_id=current_user.id,
        )
        if image_path:
            q.image_path = image_path
            db.session.commit()

        flash("Question created successfully.", "success")
        return redirect(url_for("questions.index"))

    return render_template("competitions/manage/question_form.html", form=form, question=None)


@questions_bp.route("/<int:question_id>/edit", methods=["GET", "POST"])
@login_required
def edit(question_id):
    err = staff_required()
    if err:
        return err

    question = Question.query.get_or_404(question_id)
    form = QuestionForm(obj=question)
    form.category_id.choices = [(0, "— None —")] + [
        (c.id, c.name) for c in QuestionCategory.query.filter_by(is_active=True).order_by(QuestionCategory.name.asc()).all()
    ]
    form.book_id.choices = [(0, "— None —")] + [
        (b.id, b.title) for b in Book.query.order_by(Book.title.asc()).all()
    ]

    if request.method == "GET":
        form.category_id.data = question.category_id or 0
        form.book_id.data = question.book_id or 0

    if form.validate_on_submit():
        # Handle file upload
        image_path = question.image_path
        if form.image_file.data:
            file = form.image_file.data
            ext = os.path.splitext(file.filename)[1]
            safe_filename = f"{uuid.uuid4().hex}{ext}"
            rel_dir = os.path.join("uploads", "questions")
            abs_dir = os.path.join(current_app.static_folder, rel_dir)
            os.makedirs(abs_dir, exist_ok=True)
            file.save(os.path.join(abs_dir, safe_filename))
            image_path = f"{rel_dir}/{safe_filename}"

        options_data = []
        if form.question_type.data == Question.TYPE_TRUE_FALSE:
            correct_val = request.form.get("correct_option_tf", "True")
            options_data = [
                {"option_text": "True", "is_correct": (correct_val == "True")},
                {"option_text": "False", "is_correct": (correct_val == "False")},
            ]
        else:
            options = request.form.getlist("option_text")
            correct_idxs = {int(x) for x in request.form.getlist("correct_option")}
            for idx, text in enumerate(options):
                if text.strip():
                    options_data.append({
                        "option_text": text.strip(),
                        "is_correct": idx in correct_idxs
                    })

        q = question_service.update_question(
            question_id=question_id,
            question_text=form.question_text.data,
            question_type=form.question_type.data,
            category_id=form.category_id.data,
            points=form.points.data,
            explanation=form.explanation.data,
            book_id=form.book_id.data,
            difficulty=form.difficulty.data,
            options_data=options_data,
        )
        q.image_path = image_path
        db.session.commit()

        flash("Question updated successfully.", "success")
        return redirect(url_for("questions.index"))

    return render_template("competitions/manage/question_form.html", form=form, question=question)


@questions_bp.route("/<int:question_id>/clone", methods=["POST"])
@login_required
def clone(question_id):
    err = staff_required()
    if err:
        return err

    question_service.clone_question(question_id, current_user.id)
    flash("Question cloned successfully.", "success")
    return redirect(url_for("questions.index"))


@questions_bp.route("/<int:question_id>/archive", methods=["POST"])
@login_required
def archive(question_id):
    err = staff_required()
    if err:
        return err

    question_service.archive_question(question_id)
    flash("Question archived successfully.", "success")
    return redirect(url_for("questions.index"))


@questions_bp.route("/<int:question_id>/delete", methods=["POST"])
@login_required
def delete(question_id):
    err = staff_required()
    if err:
        return err

    ok, msg = question_service.delete_question(question_id)
    if ok:
        flash(msg, "success")
    else:
        flash(msg, "warning")
    return redirect(url_for("questions.index"))


@questions_bp.route("/<int:question_id>/preview")
@login_required
def preview(question_id):
    err = staff_required()
    if err:
        return err

    question = Question.query.get_or_404(question_id)
    stats = question_service.get_question_stats(question_id)
    return render_template(
        "competitions/manage/question_preview_modal.html",
        question=question,
        stats=stats
    )


# ──────────────────────────────────────────────
# IMPORT — Upload page
# ──────────────────────────────────────────────
@questions_bp.route("/import", methods=["GET", "POST"])
@login_required
def import_questions():
    err = staff_required()
    if err:
        return err

    if request.method == "POST":
        file = request.files.get("file")
        if not file or not file.filename:
            flash("Please select an Excel file.", "warning")
            return redirect(url_for("questions.import_questions"))

        ext = os.path.splitext(file.filename)[1].lower()
        if ext not in (".xlsx", ".xls"):
            flash("Only .xlsx and .xls files are supported.", "danger")
            return redirect(url_for("questions.import_questions"))

        try:
            import openpyxl
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        except Exception as e:
            flash(f"Could not read file: {e}", "danger")
            return redirect(url_for("questions.import_questions"))

        if not rows:
            flash("The file is empty.", "warning")
            return redirect(url_for("questions.import_questions"))

        # Parse header
        header = [str(c).strip().lower() if c else "" for c in rows[0]]
        preview = []
        for i, row in enumerate(rows[1:], start=2):
            cell = dict(zip(header, row))
            item = _parse_import_row(cell, i)
            preview.append(item)

        session["import_preview"] = preview
        session["import_filename"] = file.filename
        return redirect(url_for("questions.import_preview"))

    return render_template("competitions/manage/import_questions.html")


def _parse_import_row(cell, row_num):
    """Parse one Excel row into a preview dict."""
    item = {
        "row": row_num,
        "valid": True,
        "error": None,
        "question_text": (str(cell.get("question_text") or "")).strip(),
        "question_type": (str(cell.get("question_type") or "single")).strip().lower(),
        "difficulty": (str(cell.get("difficulty") or "medium")).strip().lower(),
        "points": int(cell.get("points") or 1),
        "explanation": (str(cell.get("explanation") or "")).strip() or None,
        "category": (str(cell.get("category") or "")).strip() or None,
        "options": [],
    }

    if not item["question_text"]:
        item["valid"] = False
        item["error"] = "question_text is empty"
        return item

    VALID_TYPES = {"single", "multiple", "true_false"}
    if item["question_type"] not in VALID_TYPES:
        item["valid"] = False
        item["error"] = f"Unknown question_type '{item['question_type']}'"
        return item

    correct_raw = str(cell.get("correct_options") or "").strip()

    if item["question_type"] == "true_false":
        val = correct_raw.capitalize()
        if val not in ("True", "False"):
            item["valid"] = False
            item["error"] = "correct_options must be 'True' or 'False' for true_false type"
            return item
        item["options"] = [
            {"text": "True",  "is_correct": val == "True"},
            {"text": "False", "is_correct": val == "False"},
        ]
    else:
        try:
            correct_idxs = {int(x.strip()) for x in correct_raw.split(",") if x.strip().isdigit()}
        except Exception:
            correct_idxs = set()

        opts = []
        for n in range(1, 7):
            txt = str(cell.get(f"option_{n}") or "").strip()
            if txt:
                opts.append({"text": txt, "is_correct": n in correct_idxs})

        if len(opts) < 2:
            item["valid"] = False
            item["error"] = "At least 2 options required"
            return item

        if not any(o["is_correct"] for o in opts):
            item["valid"] = False
            item["error"] = "No correct option marked"
            return item

        item["options"] = opts

    return item


# ──────────────────────────────────────────────
# IMPORT — Preview page
# ──────────────────────────────────────────────
@questions_bp.route("/import/preview")
@login_required
def import_preview():
    err = staff_required()
    if err:
        return err

    preview = session.get("import_preview")
    if not preview:
        flash("No import data found. Please upload a file first.", "warning")
        return redirect(url_for("questions.import_questions"))

    return render_template(
        "competitions/manage/import_preview.html",
        preview=preview,
        filename=session.get("import_filename", "file.xlsx"),
    )


# ──────────────────────────────────────────────
# IMPORT — Confirm (write to DB)
# ──────────────────────────────────────────────
@questions_bp.route("/import/confirm", methods=["POST"])
@login_required
def import_confirm():
    err = staff_required()
    if err:
        return err

    preview = session.pop("import_preview", [])
    session.pop("import_filename", None)

    valid_rows = [r for r in preview if r.get("valid")]
    if not valid_rows:
        flash("No valid rows to import.", "warning")
        return redirect(url_for("questions.import_questions"))

    imported = 0
    for row in valid_rows:
        try:
            q = Question(
                question_text=row["question_text"],
                question_type=row["question_type"],
                difficulty=row["difficulty"],
                points=row["points"],
                explanation=row["explanation"],
                is_active=True,
                created_by_id=current_user.id,
            )
            db.session.add(q)
            db.session.flush()  # get q.id

            for idx, opt in enumerate(row["options"]):
                db.session.add(QuestionOption(
                    question_id=q.id,
                    option_text=opt["text"],
                    is_correct=opt["is_correct"],
                    sort_order=idx,
                ))
            imported += 1
        except Exception as e:
            db.session.rollback()
            flash(f"Error saving row {row.get('row')}: {e}", "danger")
            return redirect(url_for("questions.import_questions"))

    db.session.commit()
    flash(f"Successfully imported {imported} question{'s' if imported != 1 else ''}.", "success")
    return redirect(url_for("questions.index"))


# ──────────────────────────────────────────────
# EXPORT — Download question bank as Excel
# ──────────────────────────────────────────────
@questions_bp.route("/export")
@login_required
def export_questions():
    err = staff_required()
    if err:
        return err

    try:
        import openpyxl
    except ImportError:
        flash("openpyxl is required for export.", "danger")
        return redirect(url_for("questions.index"))

    questions = Question.query.filter_by(is_archived=False).order_by(Question.id.asc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Questions"
    ws.append([
        "question_text", "question_type", "difficulty", "points", "explanation",
        "option_1", "option_2", "option_3", "option_4", "option_5", "option_6",
        "correct_options",
    ])

    for q in questions:
        opts = sorted(q.options, key=lambda o: o.sort_order)
        opt_texts   = [o.option_text for o in opts]
        correct_idx = [str(i + 1) for i, o in enumerate(opts) if o.is_correct]
        # Pad to 6 columns
        while len(opt_texts) < 6:
            opt_texts.append("")
        ws.append([
            q.question_text,
            q.question_type,
            q.difficulty or "medium",
            q.points,
            q.explanation or "",
            *opt_texts[:6],
            ",".join(correct_idx),
        ])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="question_bank_export.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ──────────────────────────────────────────────
# TEMPLATE — Download blank import template
# ──────────────────────────────────────────────
@questions_bp.route("/template")
@login_required
def download_template():
    err = staff_required()
    if err:
        return err

    try:
        import openpyxl
    except ImportError:
        flash("openpyxl is required.", "danger")
        return redirect(url_for("questions.index"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Questions"

    headers = [
        "question_text", "question_type", "difficulty", "points", "explanation",
        "option_1", "option_2", "option_3", "option_4", "option_5", "option_6",
        "correct_options",
    ]
    ws.append(headers)

    # Style header row
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = 22

    # Example rows
    ws.append([
        "What is the capital of France?", "single", "easy", 1, "",
        "Paris", "London", "Berlin", "Rome", "", "", "1",
    ])
    ws.append([
        "Which are prime numbers?", "multiple", "medium", 2, "Prime = divisible only by 1 and itself",
        "2", "3", "4", "5", "", "", "1,2,4",
    ])
    ws.append([
        "The Earth is flat.", "true_false", "easy", 1, "",
        "", "", "", "", "", "", "False",
    ])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="question_import_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
