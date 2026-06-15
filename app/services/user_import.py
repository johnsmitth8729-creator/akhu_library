from __future__ import annotations

import re
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any

from app import db
from app.models.faculty import Faculty
from app.models.user import User
from app.utils.phone import is_valid_phone, normalize_phone
from app.utils.username import build_username_base, reserve_username


EXPECTED_HEADERS = {
    "full name": "fullname",
    "fullname": "fullname",
    "name": "fullname",
    "email": "email",
    "email address": "email",
    "phone number": "phone",
    "phone": "phone",
    "telefon": "phone",
    "group": "group",
    "guruh": "group",
}


def _normalize_email(value: str) -> str:
    return (value or "").strip().lower()


def _is_valid_email(value: str) -> bool:
    if not value or "@" not in value or len(value) > 120:
        return False
    local, _, domain = value.partition("@")
    return bool(local and domain and "." in domain)


@dataclass
class ImportRow:
    row_number: int
    fullname: str = ""
    email: str = ""
    phone: str = ""
    group: str = ""
    username: str = ""
    status: str = "valid"
    message: str = ""


@dataclass
class ImportPreview:
    faculty_id: int
    faculty_name: str
    rows: list[ImportRow] = field(default_factory=list)

    @property
    def total_rows(self) -> int:
        return len(self.rows)

    @property
    def valid_rows(self) -> list[ImportRow]:
        return [row for row in self.rows if row.status == "valid"]

    @property
    def error_rows(self) -> list[ImportRow]:
        return [row for row in self.rows if row.status != "valid"]

    @property
    def valid_count(self) -> int:
        return len(self.valid_rows)

    @property
    def error_count(self) -> int:
        return len(self.error_rows)


@dataclass
class ImportResult:
    processed: int = 0
    created: int = 0
    skipped: int = 0
    report_rows: list[dict[str, Any]] = field(default_factory=list)


def default_temp_password() -> str:
    from datetime import datetime

    year = datetime.now().year
    return f"Akhu{year}!"


def _normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _load_workbook_rows(file_bytes: bytes, filename: str) -> list[list[Any]]:
    lower_name = (filename or "").lower()

    if lower_name.endswith(".xlsx"):
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        sheet = workbook.active
        return [list(row) for row in sheet.iter_rows(values_only=True)]

    if lower_name.endswith(".xls"):
        import xlrd

        book = xlrd.open_workbook(file_contents=file_bytes)
        sheet = book.sheet_by_index(0)
        return [sheet.row_values(row_index) for row_index in range(sheet.nrows)]

    raise ValueError("Unsupported file format. Upload .xlsx or .xls files only.")


def _map_headers(header_row: list[Any]) -> dict[int, str]:
    mapping: dict[int, str] = {}
    for index, cell in enumerate(header_row):
        key = _normalize_header(cell)
        if key in EXPECTED_HEADERS:
            mapping[index] = EXPECTED_HEADERS[key]
    return mapping


def parse_excel_file(file_bytes: bytes, filename: str, faculty_id: int) -> ImportPreview:
    faculty = Faculty.query.get(faculty_id)
    if not faculty:
        raise ValueError("Selected faculty was not found.")

    raw_rows = _load_workbook_rows(file_bytes, filename)
    if not raw_rows:
        raise ValueError("The Excel file is empty.")

    header_mapping = _map_headers(raw_rows[0])
    required_fields = {"fullname", "email", "phone", "group"}
    if not required_fields.issubset(set(header_mapping.values())):
        raise ValueError(
            "Invalid Excel structure. Required columns: "
            "Full Name, Email Address, Phone Number, Group."
        )

    preview = ImportPreview(faculty_id=faculty.id, faculty_name=faculty.name)
    seen_phones: set[str] = set()
    seen_emails: set[str] = set()
    reserved_usernames: set[str] = set()
    existing_phones = {
        normalize_phone(phone)
        for phone, in db.session.query(User.phone_number).filter(User.phone_number.isnot(None)).all()
        if phone
    }
    existing_emails = {
        _normalize_email(email)
        for email, in db.session.query(User.email).all()
        if email
    }

    for offset, raw in enumerate(raw_rows[1:], start=2):
        if not any(cell is not None and str(cell).strip() for cell in raw):
            continue

        row = ImportRow(row_number=offset)
        values: dict[str, str] = {}
        for index, field_name in header_mapping.items():
            cell = raw[index] if index < len(raw) else ""
            values[field_name] = str(cell).strip() if cell is not None else ""

        row.fullname = values.get("fullname", "")
        row.email = _normalize_email(values.get("email", ""))
        row.phone = values.get("phone", "")
        row.group = values.get("group", "")

        if not row.fullname:
            row.status = "error"
            row.message = "Full name is required."
            preview.rows.append(row)
            continue

        if not row.email:
            row.status = "error"
            row.message = "Email address is required."
            preview.rows.append(row)
            continue

        if not _is_valid_email(row.email):
            row.status = "error"
            row.message = "Invalid email address."
            preview.rows.append(row)
            continue

        if row.email in seen_emails:
            row.status = "error"
            row.message = "Duplicate email in this file."
            preview.rows.append(row)
            continue

        if row.email in existing_emails:
            row.status = "error"
            row.message = "Email already exists."
            preview.rows.append(row)
            continue

        if not row.phone:
            row.status = "error"
            row.message = "Phone number is required."
            preview.rows.append(row)
            continue

        if not row.group:
            row.status = "error"
            row.message = "Group is required."
            preview.rows.append(row)
            continue

        if not is_valid_phone(row.phone):
            row.status = "error"
            row.message = "Invalid phone number."
            preview.rows.append(row)
            continue

        normalized_phone = normalize_phone(row.phone)
        if normalized_phone in seen_phones:
            row.status = "error"
            row.message = "Duplicate phone number in this file."
            preview.rows.append(row)
            continue

        if normalized_phone in existing_phones:
            row.status = "error"
            row.message = "Phone number already exists."
            preview.rows.append(row)
            continue

        seen_phones.add(normalized_phone)
        seen_emails.add(row.email)
        base_username = build_username_base(row.fullname, row.phone)
        try:
            row.username = reserve_username(base_username, reserved_usernames)
        except ValueError:
            row.status = "error"
            row.message = "Could not generate a unique username."
            preview.rows.append(row)
            continue

        row.status = "valid"
        row.message = "Ready to import"
        preview.rows.append(row)

    if not preview.rows:
        raise ValueError("No data rows found in the Excel file.")

    return preview


def preview_to_session_dict(preview: ImportPreview) -> dict[str, Any]:
    return {
        "faculty_id": preview.faculty_id,
        "faculty_name": preview.faculty_name,
        "rows": [
            {
                "row_number": row.row_number,
                "fullname": row.fullname,
                "email": row.email,
                "phone": row.phone,
                "group": row.group,
                "username": row.username,
                "status": row.status,
                "message": row.message,
            }
            for row in preview.rows
        ],
    }


def preview_from_session_dict(data: dict[str, Any]) -> ImportPreview:
    preview = ImportPreview(
        faculty_id=data["faculty_id"],
        faculty_name=data.get("faculty_name", ""),
    )
    for row in data.get("rows", []):
        preview.rows.append(
            ImportRow(
                row_number=row["row_number"],
                fullname=row["fullname"],
                email=row.get("email", ""),
                phone=row["phone"],
                group=row["group"],
                username=row["username"],
                status=row["status"],
                message=row.get("message", ""),
            )
        )
    return preview


def commit_import(preview: ImportPreview, password: str | None = None) -> ImportResult:
    faculty = Faculty.query.get_or_404(preview.faculty_id)
    temp_password = password or default_temp_password()
    result = ImportResult()

    for row in preview.rows:
        result.processed += 1

        if row.status != "valid":
            result.skipped += 1
            result.report_rows.append(
                {
                    "row": row.row_number,
                    "name": row.fullname,
                    "reason": row.message or "Skipped",
                }
            )
            continue

        normalized_phone = normalize_phone(row.phone)
        if User.query.filter_by(phone_number=normalized_phone).first():
            result.skipped += 1
            row.status = "error"
            row.message = "Phone number already exists."
            result.report_rows.append(
                {
                    "row": row.row_number,
                    "name": row.fullname,
                    "reason": row.message,
                }
            )
            continue

        if User.query.filter_by(username=row.username).first():
            result.skipped += 1
            result.report_rows.append(
                {
                    "row": row.row_number,
                    "name": row.fullname,
                    "reason": "Username already exists.",
                }
            )
            continue

        if User.query.filter_by(email=row.email).first():
            result.skipped += 1
            result.report_rows.append(
                {
                    "row": row.row_number,
                    "name": row.fullname,
                    "reason": "Email already exists.",
                }
            )
            continue

        user = User(
            fullname=row.fullname.strip(),
            username=row.username,
            email=row.email,
            phone_number=normalized_phone,
            faculty_id=faculty.id,
            faculty=faculty.name,
            group_name=row.group.strip(),
            role=User.ROLE_USER,
            email_verified=False,
        )
        user.set_password(temp_password)
        db.session.add(user)
        result.created += 1
        result.report_rows.append(
            {
                "row": row.row_number,
                "name": row.fullname,
                "reason": "Created",
            }
        )

    db.session.commit()
    return result


def build_import_report_workbook(report_rows: list[dict[str, Any]]):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Import Report"
    headers = ["Row", "Name", "Reason"]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="0B3B6F")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D9E2F1"),
        right=Side(style="thin", color="D9E2F1"),
        top=Side(style="thin", color="D9E2F1"),
        bottom=Side(style="thin", color="D9E2F1"),
    )

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for item in report_rows:
        sheet.append([item.get("row", ""), item.get("name", ""), item.get("reason", "")])

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    sheet.column_dimensions["A"].width = 10
    sheet.column_dimensions["B"].width = 32
    sheet.column_dimensions["C"].width = 40

    return workbook
